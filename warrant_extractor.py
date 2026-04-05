"""
Property Alias Warrant Data Extractor
Extracts Name, Address, City/State/Zip, and Total from warrant PDFs
and outputs them to an Excel spreadsheet.
Supports both text-based and image-based (scanned) PDFs.
"""

import os
import re
import sys
import threading
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from datetime import datetime

try:
    import pdfplumber
except ImportError:
    print("ERROR: pdfplumber is required. Install with: pip install pdfplumber")
    sys.exit(1)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)

# OCR imports (optional, used for image-based PDFs)
OCR_AVAILABLE = False
try:
    import pytesseract
    from pdf2image import convert_from_path
    OCR_AVAILABLE = True
except ImportError:
    pass


def parse_text_for_data(text):
    """
    Parse extracted text (from either pdfplumber or OCR) to find
    Name, Address, City, State, Zip, and Total.
    """
    record = {"name": "", "address": "", "city": "", "state": "", "zip_code": "", "total": ""}

    lines = [l.strip() for l in text.split("\n") if l.strip()]

    # Find name/address/city after "collect forthwith from"
    for i, line in enumerate(lines):
        if "forthwith" in line.lower() and "from" in line.lower():
            remaining = [l for l in lines[i + 1:] if l.strip()]
            info_lines = []
            for r in remaining:
                if "sum of" in r.lower() or "see below" in r.lower():
                    break
                if r.strip():
                    info_lines.append(r.strip())
                if len(info_lines) == 3:
                    break

            if len(info_lines) >= 1:
                record["name"] = info_lines[0]
            if len(info_lines) >= 2:
                record["address"] = info_lines[1]
            if len(info_lines) >= 3:
                csz_full = info_lines[2]
                csz_parts = csz_full.rsplit(" ", 2)
                if len(csz_parts) == 3:
                    record["city"] = csz_parts[0]
                    record["state"] = csz_parts[1]
                    record["zip_code"] = csz_parts[2]
                else:
                    record["city"] = csz_full
            break

    # Find total on or near "BALANCE DUE" line
    for i, line in enumerate(lines):
        if "balance due" in line.lower():
            amounts = re.findall(r'[\d,]+\.\d{2}', line)
            if amounts:
                record["total"] = amounts[-1]
            else:
                for nearby in lines[i + 1: i + 3]:
                    amounts = re.findall(r'[\d,]+\.\d{2}', nearby)
                    if amounts:
                        record["total"] = amounts[-1]
                        break
            break

    return record


def extract_warrant_data(pdf_path, status_callback=None):
    """
    Extract data from each page of a warrant PDF.
    Uses pdfplumber for text-based PDFs, falls back to OCR for image-based.
    """
    results = []

    with pdfplumber.open(pdf_path) as pdf:
        total_pages = len(pdf.pages)
        is_image_pdf = False

        # Check first page to see if it has text
        first_page_words = pdf.pages[0].extract_words() if total_pages > 0 else []
        if not first_page_words:
            is_image_pdf = True

        if not is_image_pdf:
            # Text-based PDF - use pdfplumber
            for page_num, page in enumerate(pdf.pages, start=1):
                record = {
                    "source_file": os.path.basename(pdf_path),
                    "page": page_num,
                    "name": "", "address": "", "city": "",
                    "state": "", "zip_code": "", "total": "",
                }

                words = page.extract_words()
                if words:
                    lines_dict = {}
                    for w in words:
                        top_rounded = round(w["top"], 0)
                        if top_rounded not in lines_dict:
                            lines_dict[top_rounded] = []
                        lines_dict[top_rounded].append(w)

                    sorted_tops = sorted(lines_dict.keys())

                    marker_idx = None
                    for i, top in enumerate(sorted_tops):
                        line_text = " ".join(w["text"] for w in lines_dict[top])
                        if "forthwith" in line_text.lower():
                            marker_idx = i
                            break

                    if marker_idx is not None:
                        info_lines = []
                        for j in range(marker_idx + 1, len(sorted_tops)):
                            top = sorted_tops[j]
                            line_words = sorted(lines_dict[top], key=lambda w: w["x0"])
                            if line_words and line_words[0]["x0"] >= 55:
                                line_text = " ".join(w["text"] for w in line_words).strip()
                                if line_text and "sum of" not in line_text.lower():
                                    info_lines.append(line_text)
                                if len(info_lines) == 3:
                                    break
                            elif info_lines:
                                break

                        if len(info_lines) >= 1:
                            record["name"] = info_lines[0]
                        if len(info_lines) >= 2:
                            record["address"] = info_lines[1]
                        if len(info_lines) >= 3:
                            csz_full = info_lines[2]
                            csz_parts = csz_full.rsplit(" ", 2)
                            if len(csz_parts) == 3:
                                record["city"] = csz_parts[0]
                                record["state"] = csz_parts[1]
                                record["zip_code"] = csz_parts[2]
                            else:
                                record["city"] = csz_full

                    for i, top in enumerate(sorted_tops):
                        line_text = " ".join(w["text"] for w in lines_dict[top])
                        if "balance due" in line_text.lower():
                            nearby_words = []
                            for nearby_top in sorted_tops:
                                if abs(nearby_top - top) <= 5:
                                    nearby_words.extend(lines_dict[nearby_top])
                            nearby_words.sort(key=lambda w: w["x0"], reverse=True)
                            for tw in nearby_words:
                                text = tw["text"].replace(",", "").replace("$", "")
                                try:
                                    float(text)
                                    record["total"] = tw["text"]
                                    break
                                except ValueError:
                                    continue
                            break

                # Fallback: try plain text extraction
                if not record["name"] or not record["total"]:
                    text = page.extract_text() or ""
                    if text:
                        parsed = parse_text_for_data(text)
                        if not record["name"]:
                            record["name"] = parsed["name"]
                            record["address"] = parsed["address"]
                            record["city"] = parsed["city"]
                            record["state"] = parsed["state"]
                            record["zip_code"] = parsed["zip_code"]
                        if not record["total"]:
                            record["total"] = parsed["total"]

                results.append(record)

        elif is_image_pdf and OCR_AVAILABLE:
            # Image-based PDF - use OCR
            if status_callback:
                status_callback("Converting PDF pages to images for OCR...")

            images = convert_from_path(pdf_path, dpi=300)

            for page_num, img in enumerate(images, start=1):
                record = {
                    "source_file": os.path.basename(pdf_path),
                    "page": page_num,
                    "name": "", "address": "", "city": "",
                    "state": "", "zip_code": "", "total": "",
                }

                if status_callback:
                    status_callback(f"OCR processing page {page_num}/{len(images)}...")

                text = pytesseract.image_to_string(img)
                parsed = parse_text_for_data(text)
                record["name"] = parsed["name"]
                record["address"] = parsed["address"]
                record["city"] = parsed["city"]
                record["state"] = parsed["state"]
                record["zip_code"] = parsed["zip_code"]
                record["total"] = parsed["total"]

                results.append(record)

        elif is_image_pdf and not OCR_AVAILABLE:
            for page_num in range(1, total_pages + 1):
                results.append({
                    "source_file": os.path.basename(pdf_path),
                    "page": page_num,
                    "name": "OCR NOT AVAILABLE - install pytesseract and pdf2image",
                    "address": "", "city": "", "state": "", "zip_code": "", "total": "",
                })

    return results


def write_to_excel(all_records, output_path):
    """Write extracted records to a formatted Excel spreadsheet."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Warrant Data"

    header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    headers = ["Name", "Address", "City", "State", "Zip Code", "Total", "Source File", "Page"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    data_font = Font(name="Calibri", size=11)
    for row_idx, record in enumerate(all_records, start=2):
        values = [
            record["name"], record["address"], record["city"],
            record["state"], record["zip_code"], record["total"],
            record["source_file"], record["page"],
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.border = thin_border

    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 4, 50)

    wb.save(output_path)


class WarrantExtractorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Warrant Data Extractor")
        self.root.geometry("700x500")
        self.root.resizable(True, True)
        self.root.minsize(600, 400)
        self.pdf_files = []
        self.build_ui()

    def build_ui(self):
        main_frame = ttk.Frame(self.root, padding=20)
        main_frame.pack(fill=tk.BOTH, expand=True)

        title_label = ttk.Label(
            main_frame, text="Property Alias Warrant Extractor",
            font=("Calibri", 16, "bold"),
        )
        title_label.pack(pady=(0, 5))

        subtitle_label = ttk.Label(
            main_frame,
            text="Extract Name, Address, City, State, Zip, and Total from warrant PDFs",
            font=("Calibri", 10),
        )
        subtitle_label.pack(pady=(0, 15))

        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(fill=tk.X, pady=(0, 10))

        self.select_btn = ttk.Button(btn_frame, text="Select PDF Files", command=self.select_files)
        self.select_btn.pack(side=tk.LEFT, padx=(0, 5))

        self.folder_btn = ttk.Button(btn_frame, text="Select Folder", command=self.select_folder)
        self.folder_btn.pack(side=tk.LEFT, padx=(0, 5))

        self.clear_btn = ttk.Button(btn_frame, text="Clear List", command=self.clear_files)
        self.clear_btn.pack(side=tk.LEFT)

        list_frame = ttk.LabelFrame(main_frame, text="PDF Files", padding=5)
        list_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 10))

        self.file_listbox = tk.Listbox(list_frame, font=("Calibri", 10))
        scrollbar = ttk.Scrollbar(list_frame, orient=tk.VERTICAL, command=self.file_listbox.yview)
        self.file_listbox.configure(yscrollcommand=scrollbar.set)
        self.file_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        self.status_var = tk.StringVar(value="Select PDF files or a folder to begin.")
        status_label = ttk.Label(main_frame, textvariable=self.status_var, font=("Calibri", 10))
        status_label.pack(fill=tk.X, pady=(0, 5))

        self.progress = ttk.Progressbar(main_frame, mode="determinate")
        self.progress.pack(fill=tk.X, pady=(0, 10))

        self.extract_btn = ttk.Button(
            main_frame, text="Extract Data to Spreadsheet", command=self.run_extraction,
        )
        self.extract_btn.pack(pady=(0, 5))

    def select_files(self):
        files = filedialog.askopenfilenames(
            title="Select PDF Files",
            filetypes=[("PDF Files", "*.pdf"), ("All Files", "*.*")],
        )
        if files:
            for f in files:
                if f not in self.pdf_files:
                    self.pdf_files.append(f)
                    self.file_listbox.insert(tk.END, os.path.basename(f))
            self.status_var.set(f"{len(self.pdf_files)} file(s) selected.")

    def select_folder(self):
        folder = filedialog.askdirectory(title="Select Folder Containing PDFs")
        if folder:
            count = 0
            for fname in sorted(os.listdir(folder)):
                if fname.lower().endswith(".pdf"):
                    full_path = os.path.join(folder, fname)
                    if full_path not in self.pdf_files:
                        self.pdf_files.append(full_path)
                        self.file_listbox.insert(tk.END, fname)
                        count += 1
            self.status_var.set(f"{len(self.pdf_files)} file(s) selected. ({count} added from folder)")

    def clear_files(self):
        self.pdf_files.clear()
        self.file_listbox.delete(0, tk.END)
        self.progress["value"] = 0
        self.status_var.set("List cleared. Select PDF files or a folder to begin.")

    def run_extraction(self):
        if not self.pdf_files:
            messagebox.showwarning("No Files", "Please select PDF files first.")
            return

        output_path = filedialog.asksaveasfilename(
            title="Save Spreadsheet As",
            defaultextension=".xlsx",
            initialfile=f"warrant_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            filetypes=[("Excel Files", "*.xlsx"), ("All Files", "*.*")],
        )
        if not output_path:
            return

        self.extract_btn.configure(state=tk.DISABLED)
        self.select_btn.configure(state=tk.DISABLED)
        self.folder_btn.configure(state=tk.DISABLED)

        thread = threading.Thread(target=self.do_extraction, args=(output_path,), daemon=True)
        thread.start()

    def update_status(self, msg):
        self.root.after(0, lambda: self.status_var.set(msg))

    def do_extraction(self, output_path):
        all_records = []
        total_files = len(self.pdf_files)

        for idx, pdf_path in enumerate(self.pdf_files):
            self.update_status(f"Processing ({idx+1}/{total_files}): {os.path.basename(pdf_path)}")

            try:
                records = extract_warrant_data(pdf_path, status_callback=self.update_status)
                all_records.extend(records)
            except Exception as e:
                self.root.after(
                    0, lambda p=pdf_path, err=e: messagebox.showerror(
                        "Error", f"Error processing {os.path.basename(p)}:\n{err}"
                    ),
                )

            progress_val = ((idx + 1) / total_files) * 100
            self.root.after(0, lambda v=progress_val: self.progress.configure(value=v))

        if all_records:
            try:
                write_to_excel(all_records, output_path)
                self.update_status(
                    f"Done! Extracted {len(all_records)} records to {os.path.basename(output_path)}"
                )
                self.root.after(
                    0, lambda: messagebox.showinfo(
                        "Success",
                        f"Extracted {len(all_records)} records from {total_files} file(s).\n\nSaved to:\n{output_path}",
                    ),
                )
            except Exception as e:
                self.root.after(
                    0, lambda: messagebox.showerror("Error", f"Error saving spreadsheet:\n{e}")
                )
        else:
            self.root.after(
                0, lambda: messagebox.showwarning("No Data", "No records were extracted.")
            )

        self.root.after(0, lambda: self.extract_btn.configure(state=tk.NORMAL))
        self.root.after(0, lambda: self.select_btn.configure(state=tk.NORMAL))
        self.root.after(0, lambda: self.folder_btn.configure(state=tk.NORMAL))


def main():
    root = tk.Tk()
    app = WarrantExtractorApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()